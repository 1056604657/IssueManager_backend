from dvadmin.system.models import Users, JiraProject, JiraIssue, IssueComment
from dvadmin.system.views.user import UserSerializer

from dvadmin.utils.serializers import CustomModelSerializer

from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin.utils.json_response import DetailResponse, SuccessResponse, ErrorResponse
from rest_framework.decorators import action
from dingtalkchatbot.chatbot import DingtalkChatbot
from enum import Enum
from datetime import datetime, timedelta
from django.db.models import Count, DateField, Case, When, F
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl import Workbook

from apscheduler.schedulers.background import BackgroundScheduler
from django_apscheduler.jobstores import DjangoJobStore, register_events, register_job


class JiraProjectSerializer(CustomModelSerializer):
    class Meta:
        model = JiraProject
        fields = '__all__'


class IssueCommentSerializer(CustomModelSerializer):
    author = UserSerializer()

    class Meta:
        model = IssueComment
        fields = '__all__'


class AllIssueSerializer(CustomModelSerializer):
    project = JiraProjectSerializer(many=False)
    assigned = UserSerializer(many=False)
    class Meta:
        model = JiraIssue
        fields = '__all__'

class JiraIssueSerializer(CustomModelSerializer):
    class Meta:
        model = JiraIssue
        fields = '__all__'


class JiraViewSet(CustomModelViewSet):
    queryset = JiraProject.objects.all()
    serializer_class = JiraProjectSerializer

    @action(methods=['GET'], detail=False)
    def get_jira_project_stages(self, request):
        stages = JiraProject._meta.get_field('stage').choices
        stages = [{"id": id, "name": name} for id, name in stages]
        return SuccessResponse(stages)

    @action(methods=['GET'], detail=False)
    def get_jira_project_status(self, request):
        status = JiraProject._meta.get_field('status').choices
        status = [{"id": id, "name": name} for id, name in status]
        return SuccessResponse(status)

    @action(methods=['GET'], detail=False)
    # 获取项目列表
    def get_project_list_page(self, request):
        queryset = JiraProject.objects.all()
        if request.query_params.get('name'):
            queryset = queryset.filter(name__icontains=request.query_params.get('name'))
        serializer = JiraProjectSerializer(queryset, many=True)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['POST'], detail=False)
    # 编辑项目信息
    def update_project(self, request):
        data = request.data
        data1 = {
            'name': data.get('name'),
            'key': data.get('key'),
            'description': data.get('description'),
            'manager_id': data.get('manager'),
            'ding_webhook': data.get('ding_webhook'),
            'stage': data.get('stage'),
            'status': data.get('status')
        }
        JiraProject.objects.filter(id=data.get('id')).update(**data1)
        return DetailResponse(msg='更新成功')

    @action(methods=['POST'], detail=False)
    # 删除项目
    def delete_project(self, request):
        data = request.data
        count = JiraIssue.objects.filter(project=data.get('id')).count()
        if count:
            return ErrorResponse(msg='项目下存在issue,不能删除')
        JiraProject.objects.filter(id=data.get('id')).delete()
        return DetailResponse(msg='更新成功')

    @action(methods=['GET'], detail=False)
    # 获取所有项目列表
    def get_project_list(self, request):
        queryset = JiraProject.objects.all()
        serializer = JiraProjectSerializer(queryset, many=True)
        data = serializer.data
        return DetailResponse(data=data)

    @action(methods=['GET'], detail=False)
    def get_all_issue(self, request):
        queryset = JiraIssue.objects.all()
        data = request.query_params
        if data.get('project_id'):
            queryset = queryset.filter(project_id=data.get('project_id'))
        if data.get('name') is not None:
            queryset = queryset.filter(name__icontains=data.get('name'))
        if data.get('status'):
            queryset = queryset.filter(status=data.get('status'))
        if data.get('out_date') is not None:
            queryset = queryset.filter(status=1, deadline__isnull=False, deadline__lte=datetime.now())
        queryset = queryset.select_related('project').select_related('assigned')
        serializer = AllIssueSerializer(queryset, many=True)
        print(serializer.data)
        if data.get('export'):
            headers = ['所属项目', '标题', '标识号', '类型', '状态', '优先级', '延期', '解决结果', '来源', '指派给',
                       '经办人', '报告人', '创建时间', '更新时间', '到期时间', '解决时间',
                       '预期工时', '实际工时', '问题原因', '解决方法']
            result = []
            for item in serializer.data:
                row = [
                    item['project']['name'],
                    item['name'],
                    item['signal_number'],
                    dict(JiraIssue._meta.get_field('type').choices).get(item['type']),
                    '处理中' if item['status'] == 1 and item['pending_datetime'] else '待处理' if item['status'] == 1 else '已完成',
                    dict(JiraIssue._meta.get_field('priority').choices).get(item['priority']),
                    '是' if item['status'] == 1 and item['deadline'] and item['deadline'] < datetime.now() else '否',
                    '已解决' if item['resolve_datetime'] else '未解决',
                    dict(JiraIssue._meta.get_field('source').choices).get(item['source']),
                    item['assigned']['name'],
                    item['creator_name'],
                    item['modifier_name'],
                    item['create_datetime'],
                    item['update_datetime'],
                    item['deadline'],
                    item['resolve_datetime'],
                    item['expected_hours'],
                    item['actual_hours'],
                    item['question_reason'],
                    item['resolve_method']
                ]
                result.append(row)
            response = export_excel(result, headers, filename='example.xlsx')
            return response
        else:
            data = serializer.data
            return DetailResponse(data=data)

    @action(methods=['GET'], detail=False)
    # 获取个人issue列表
    def get_my_issue(self, request):
        user = request.user.id
        status = request.query_params.get('status')
        queryset = JiraIssue.objects.filter(assigned=user)
        if status:
            queryset = queryset.filter(status=status)
        serializer = JiraIssueSerializer(queryset, many=True)
        data = serializer.data
        return DetailResponse(data=data)

    @action(methods=['POST'], detail=False)
    def get_issue_detail(self, request):
        data = request.data
        issue_id = data.get('id')
        issue = JiraIssue.objects.get(id=issue_id)
        serializer = JiraIssueSerializer(issue)
        serialized_data = serializer.data
        serialized_data['assigned_name'] = '-'
        if serializer.data.get('assigned'):
            user = Users.objects.get(id=serializer.data.get('assigned'))
            if user:
                serialized_data['assigned_name'] = user.name
        serialized_data['project_name'] = '-'
        if serializer.data.get('project'):
            project = JiraProject.objects.get(id=serializer.data.get('project'))
            if project:
                serialized_data['project_name'] = project.name
        return DetailResponse(data=serialized_data)

    @action(methods=['GET'], detail=False)
    # 获取请求中projectid和status的相应issue的所有字段
    def get_issue_list(self, request):
        queryset = JiraIssue.objects.filter(project_id=request.query_params.get('project_id'))
        status = request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        serializer = JiraIssueSerializer(queryset, many=True)
        serialized_data = serializer.data
        return DetailResponse(data=serialized_data)

    @action(methods=['GET'], detail=False)
    # 获取所有用户的id和name
    def get_jira_user(self, request):
        queryset = Users.objects.all()
        serializer = UserSerializer(queryset, many=True)
        serialized_data = serializer.data
        processed_data = [{'id': user['id'], 'name': user['name']} for user in serialized_data]
        return DetailResponse(data=processed_data)

    @action(methods=['POST'], detail=False)
    def add_issue(self, request):
        data = request.data
        project_id = data.get('project_id')
        project = JiraProject.objects.get(id=project_id)
        count = JiraIssue.objects.filter(project_id=project_id).count()
        count = count + 1
        data['signal_number'] = str(project.key) + '-' + str(count)
        data['creator_id'] = request.user.id
        data['modifier'] = request.user.id
        data['status'] = 1
        JiraIssue.objects.create(**data)
        assigned_user = Users.objects.get(id=data['assigned_id'])
        user_serializer = UserSerializer(assigned_user)
        assigned_name = user_serializer.data.get('name')
        data['assigned_name'] = assigned_name
        project_id2 = JiraProject.objects.get(id=data['project_id'])
        Project_Serializer = JiraProjectSerializer(project_id2)
        project_name = Project_Serializer.data.get('name')
        # 获取用户手机号码
        assigned_mobile = user_serializer.data.get('mobile')

        # 在发送钉钉消息时添加 mobiles 参数
        send_dingtalk_message(project.ding_webhook,
                              '新的现场问题已创建：来自于项目' + project_name + '， issue标题为: ' + data[
                                  "name"] + '，请登录BMEIM确认: http://im.bmetech.com',
                              mobiles=[assigned_mobile])
        return DetailResponse(data='创建成功')

    @action(methods=['POST'], detail=False)
    def update_issue(self, request):
        data = request.data
        issue = JiraIssue.objects.filter(id=data.get('id'))
        if not issue:
            return ErrorResponse(msg='issue不存在')
        data.pop('id', None)
        data['modifier'] = request.user.id
        data['update_datetime'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        issue.update(**data)
        return DetailResponse(data='保存成功')

    @action(methods=['POST'], detail=False)
    def resolve_issue(self, request):
        data = request.data
        issue_id = data.get('id')
        issue = JiraIssue.objects.filter(id=issue_id)
        if not issue:
            return ErrorResponse(msg='issue不存在')
        update_data = {
                       'resolve_datetime': data.get('resolve_datetime'), 'modifier': request.user.id, 'status': 2,
                       'update_datetime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                       'actual_hours': data.get('actual_hours'),
                       'question_reason': data.get('comment'),
                       'resolve_method': data.get('solution')
                       }
        JiraIssue.objects.filter(id=issue_id).update(**update_data)
        issue = JiraIssue.objects.get(id=issue_id)
        issue_serializer = JiraIssueSerializer(issue)
        project = JiraProject.objects.get(id=issue_serializer.data.get('project'))
        project_serialize = JiraProjectSerializer(project)
        if project_serialize.data.get('ding_webhook'):
            user = Users.objects.get(id=request.user.id)
            user_serializer = UserSerializer(user)
            send_dingtalk_message(project.ding_webhook,
                                  user_serializer.data.get('name') + '解决了issue:' + issue_serializer.data.get(
                                      'signal_number') + ' ' + issue_serializer.data.get('name'), mobiles=None)
        return DetailResponse(data='保存成功')

    @action(methods=['POST'], detail=False)
    def confirm_issue(self, request):
        data = request.data
        issue = JiraIssue.objects.filter(id=data.get('id'))
        if not issue:
            return ErrorResponse(msg='issue不存在')
        params = {
            'pending_datetime': data.get('pending'),
            'expected_hours': data.get('expected_hours'),
            'update_datetime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'modifier': request.user.id
        }
        JiraIssue.objects.filter(id=data.get('id')).update(**params)
        return DetailResponse(msg='保存成功')

    @action(methods=['GET'], detail=False)
    def issue_analysis(self, request):
        issue_count = JiraIssue.objects.count()
        unpending_count = JiraIssue.objects.filter(status=1, pending_datetime__isnull=True).count()
        pending_count = JiraIssue.objects.filter(status=1, pending_datetime__isnull=False).count()
        resolve_count = JiraIssue.objects.filter(status=2).count()
        seven_days_ago = datetime.now() - timedelta(days=7)
        daily_counts = JiraIssue.objects.filter(
            create_datetime__gte=seven_days_ago
        ).values('create_datetime__date').annotate(
            created_count=Coalesce(Count('id'), 0),
            resolved_count=Coalesce(Count(Case(When(resolve_datetime__date=F('create_datetime__date'), then=1))), 0)
        ).order_by('create_datetime__date')
        daily_data = []
        current_date = seven_days_ago.date()
        for entry in daily_counts:
            while entry['create_datetime__date'] > current_date:
                daily_data.append({'date': current_date, 'created_count': 0, 'resolved_count': 0})
                current_date += timedelta(days=1)
            daily_data.append({
                'date': entry['create_datetime__date'],
                'created_count': entry['created_count'],
                'resolved_count': entry['resolved_count'],
            })
            current_date += timedelta(days=1)

        while current_date <= datetime.now().date():
            daily_data.append({'date': current_date, 'created_count': 0, 'resolved_count': 0})
            current_date += timedelta(days=1)

        projects = (JiraIssue.objects.filter(resolve_datetime__isnull=True).values('project__name')
                    .annotate(unresolve_count=Count('id')).order_by('project__name'))
        project_data = []
        for entry in projects:
            project_data.append({
                'project_name': entry['project__name'],
                'unresolve_count': entry['unresolve_count']
            })
        data = {
            'issue_count': issue_count,
            'unpending_count': unpending_count,
            'pending_count': pending_count,
            'resolve_count': resolve_count,
            'daily_count': daily_data,
            'unresolved_project_count': project_data
        }
        return DetailResponse(data=data)


def send_dingtalk_message(webhook, msg, mobiles):
    try:
        bot = DingtalkChatbot(webhook)
        if mobiles:
            bot.send_text(msg=msg, at_mobiles=mobiles)
        else:
            bot.send_text(msg=msg)
        print("Dingtalk message sent successfully!")
    except Exception as e:
        print(f"Error sending Dingtalk message: {str(e)}")


scheduler = BackgroundScheduler()
scheduler.add_jobstore(DjangoJobStore(), 'default')


# 每天早上9点执行这个任务
@register_job(scheduler, 'cron', id='jira_check_outdated', hour=9, replace_existing=True)
def jira_check_outdated():
    current_time = datetime.now()
    projects = JiraProject.objects.filter(ding_webhook__isnull=False)
    for project in projects:
        issue_by_assigned = {}
        issues = (
            JiraIssue.objects.filter(deadline__gt=current_time, status=1, project=project.id, assigned__isnull=False).
            values('id', 'name', 'assigned', 'assigned__name', 'assigned__mobile', 'assigned_id'))
        user_map = {}
        for issue in issues:
            user_map[issue.get('assigned_id')] = {
                'name': issue.get('assigned__name'),
                'mobile': issue.get('assigned__mobile')
            }
        for issue in issues:
            assigned_id = issue.get('assigned_id')
            if assigned_id in issue_by_assigned:
                issue_by_assigned[assigned_id] += 1
            else:
                issue_by_assigned[assigned_id] = 1
        for assigned_id, issue_count in issue_by_assigned.items():
            name = user_map[assigned_id].get('name')
            text = f'{name},项目:{project.name}有{issue_count}个issue已过期,请即使处理'
            if user_map[assigned_id].get('mobile'):
                send_dingtalk_message(project.ding_webhook,
                                      text,
                                      mobiles=[user_map[assigned_id].get('mobile')])
            else:
                send_dingtalk_message(project.ding_webhook,
                                      text,
                                      mobiles=None)
    pass


register_events(scheduler)
scheduler.start()


def export_excel(data, headers, filename='exported_data.xlsx'):
    wb = Workbook()
    ws = wb.active
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(ord('A') + col_num - 1)
        cell = ws[f'{col_letter}1']
        cell.value = header

    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            col_letter = chr(ord('A') + col_num - 1)
            cell = ws[f'{col_letter}{row_num}']
            cell.value = value

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response
